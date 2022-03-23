import openpyxl
import os
from common.handler_path import ab_path

class Open_Excel:
    # 初始化加载文件信息,文件路径，sheet名称
    def __init__(self, file_name, sheet_name=None):
        self.file_name = ab_path(file_name)
        self.sheet_name = sheet_name

    # 读取excel文件方法
    def read_excel(self):
        # 创建文件对象
        book = openpyxl.load_workbook(self.file_name)
        if self.sheet_name is None:
            # 创建sheet页对象
            sheet = book.active
        else:
            sheet = book[self.sheet_name]
        # 以元组获取所有行数据
        rows = tuple(sheet.rows)
        # 存储首行数据
        data_title = []
        for title in rows[0]:
            data_title.append(title.value)

        data = []
        for i in range(1, len(rows)):
            data1 = []
            for s in rows[i]:
                data1.append(s.value)
            data.append(dict(zip(data_title, data1)))
        book.close()
        # 返回列表嵌套字典数据
        return data

    # 写入excel数据
    def write_excel(self, rows, colum, resule):
        book = openpyxl.load_workbook(self.file_name)
        if self.sheet_name is None:
            sheet = book.active
        else:
            sheet = book[self.sheet_name]

        sheet.cell(rows, colum).value = resule
        # sheet.cell(rows,sheet.max_column-1).value=resule
        book.save(self.file_name)
        book.close()
